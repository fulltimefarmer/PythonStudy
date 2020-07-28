from openpyxl import load_workbook

wb = load_workbook('../file/students.xlsx')

print(wb.sheetnames)

sheet = wb.worksheets[0]
print('Max row: %d' % sheet.max_row)
print('Max column: %d' % sheet.max_column)
print()

for row in sheet.rows:
    for cell in row:
        print(cell.value, end='\t\t')
    print()


