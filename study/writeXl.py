import datetime
from random import choice
from openpyxl import Workbook
from openpyxl import utils

wb = Workbook()

ws = wb.create_sheet("Sheet1", 0)

columns = []

for i in range(300):
    columns.append('column%d' % i)

ws.append(columns)

print(datetime.datetime.now().strftime("%H:%M:%S"))
for i in range(10):
    row = []
    for j in range(300):
        row.append(utils.get_column_letter(choice(range(1, 50))))
        #TITLE = 'No %d' % i
        #TIME = datetime.datetime.now().strftime("%H:%M:%S")
        #A_Z = utils.get_column_letter(choice(range(1, 50)))
    ws.append(row)
print(datetime.datetime.now().strftime("%H:%M:%S"))

wb.save('../file/test.xlsx')