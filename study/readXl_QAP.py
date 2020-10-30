from openpyxl import load_workbook

# wb_key_look_mapping = load_workbook('../file/key_look_mapping.xlsx')
# ws_key_look_mapping = wb_key_look_mapping.worksheets[0]
# for row in ws_key_look_mapping.rows:
#     dict_data = {'concept': row[1].value, 'key': row[2].value}
#     dict_key_look_mapping[row[0].value] = dict_data



wb_ae_qap = load_workbook('../file/initial_qap.xlsx')
ws_ae_qap = wb_ae_qap.worksheets[0]

max_column = ws_ae_qap.max_column

dict_key_look_mapping = {}

for i in range(1, max_column):
    cell = ws_ae_qap.cell(row=1, column=i)
    val = cell.value.strip()
    if '-APS' in val:
        break
    if val.startswith('BEACON') and val.endswith('KEY'):
        dict_key_look_mapping['BEACON'] = val
    if val.startswith('NIKE RUN') and val.endswith('KEY'):
        dict_key_look_mapping['NIKE RUN'] = val
    if val.startswith('NIKE SPORT CREATE') and val.endswith('KEY'):
        dict_key_look_mapping['NIKE SPORT CREATE'] = val
    if val.startswith('NIKE SPORT CATALYZE') and val.endswith('KEY'):
        dict_key_look_mapping['NIKE SPORT CATALYZE'] = val
    if val.startswith('NIKE KIDS') and val.endswith('KEY'):
        dict_key_look_mapping['NIKE KIDS'] = val
    if val.startswith('KL SIS') and val.endswith('KEY'):
        dict_key_look_mapping['KL SIS'] = val
    if val.startswith('KL L1L2') and val.endswith('KEY'):
        dict_key_look_mapping['KL L1L2'] = val
    if val.startswith('HOOPS') and val.endswith('KEY'):
        dict_key_look_mapping['HOOP'] = val
    if val.startswith('JORDAN') and val.endswith('KEY'):
        dict_key_look_mapping['JORDAN'] = val
    if val.startswith('KICKS LOUNGE') and val.endswith('KEY'):
        dict_key_look_mapping['KICKS LOUNGE/KL'] = val
    if val.startswith('SIS BB ') and val.endswith('KEY'):
        dict_key_look_mapping['BB'] = val
    if val.startswith('SIS RN') and val.endswith('KEY'):
        dict_key_look_mapping['RN'] = val
    if val.startswith('YOUNG ATHLETES') and val.endswith('KEY'):
        dict_key_look_mapping['YOUNG ATHLETES'] = val

for k in dict_key_look_mapping.keys():
    print('%s : %s' % (k, dict_key_look_mapping[k]))




